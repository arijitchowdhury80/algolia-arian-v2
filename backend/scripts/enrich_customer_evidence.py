"""Enrich customer evidence database from CustomerEvidence-Algolia.xlsx.

This script:
1. Normalizes industry values (maps messy variants → canonical names)
2. Enriches algolia_customers with ARR tier from vertical tabs
3. Adds partner cross-reference from Adobe tab
4. Merges Recommend Quotes into algolia_quotes
5. Merges Neural Search CX into algolia_customers

Idempotent — safe to run multiple times.
"""

from __future__ import annotations

import os
import sys
from datetime import UTC, datetime
from pathlib import Path

import openpyxl
import psycopg2

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

# docs/ moved to the repo root in the 2026-07-28 restructure.
EXCEL_PATH = (
    Path(__file__).resolve().parents[2] / "docs" / "data" / "CustomerEvidence-Algolia.xlsx"
)
DB_URL = os.environ.get("DATABASE_URL_SYNC", "postgresql://prism:prism_dev_password@localhost:5432/prism")

# ---------------------------------------------------------------------------
# Industry normalization map
# ---------------------------------------------------------------------------

INDUSTRY_MAP: dict[str, str] = {
    # Fashion cluster
    "Fashion": "Fashion",
    "E-Commerce / Clothing": "Fashion",
    "E-Comm / Clothing": "Fashion",
    "E-Commerce / Clothing & Accessories": "Fashion",
    "Ecommerce / Clothing": "Fashion",
    "Ecommerce / Apparel": "Fashion",
    "E-Commerce / Apparel": "Fashion",
    "Ecommerce / Clothing & Accessories": "Fashion",
    "Ecommerce / Retail / Fashion": "Fashion",
    "B2C/ Clothing and Clothing Accessories Retailers": "Fashion",
    "E-Commerce / Shoes": "Fashion",
    "E-Commerce / Sneakers": "Fashion",

    # Luxury
    "Luxury": "Luxury",
    "E-Commerce / Watches": "Luxury",
    "E-Commerce / Jewelry": "Luxury",
    "Ecommerce / Jewelry": "Luxury",
    "E-Comm / Jewelry": "Luxury",
    "E-Commerce / Cosmetics + Skin Care": "Luxury",

    # Grocery
    "Grocery": "Grocery",
    "E-Commerce / Grocery": "Grocery",
    "Ecommerce / Grocery": "Grocery",
    "E-Comm / Grocery": "Grocery",

    # E-commerce (general)
    "E-commerce Retail": "E-commerce Retail",
    "E-Comm": "E-commerce Retail",
    "E-Commerce": "E-commerce Retail",
    "Ecommerce/ Retail": "E-commerce Retail",
    "E-Commerce / Retail Group": "E-commerce Retail",
    "E-Commerce / Electronics": "E-commerce Retail",
    "E-Commerce / Furniture": "E-commerce Retail",
    "Ecommerce / Furniture": "E-commerce Retail",
    "E-Commerce / Baby Goods": "E-commerce Retail",
    "E-Commerce / Tools": "E-commerce Retail",
    "E-Commerce / Musical Instruments": "E-commerce Retail",
    "E-Comm / Sporting Goods": "E-commerce Retail",
    "E-Comm / Pet Supplies": "E-commerce Retail",
    "Ecommerce / Online Pharmacy": "Healthcare",
    "E-Commerce / Online Pharmacy": "Healthcare",
    "E-Comm / Online Pharmacy": "Healthcare",

    # Marketplace
    "E-commerce - Marketplace": "Marketplace",
    "E-Comm / Marketplace": "Marketplace",
    "Marketplace": "Marketplace",

    # B2B
    "B2B SaaS/PaaS": "B2B SaaS",
    "B2B E-Commerce": "B2B E-Commerce",
    "B2B E-Comm": "B2B E-Commerce",

    # Media
    "Media/Content": "Media & Content",
    "Media": "Media & Content",
    "Media / Streaming": "Media & Content",
    "Media / Online Content": "Media & Content",

    # Travel
    "Travel": "Travel & Hospitality",

    # Financial Services
    "Financial Services": "Financial Services",
    "Misc / Bank": "Financial Services",
    "Misc / Insurance": "Financial Services",

    # SaaS
    "SaaS": "SaaS",
    "SaaS / OEM": "SaaS",
    "SaaS / E-Learning": "SaaS",
    "Internal Search": "SaaS",

    # Other
    "Misc": "Other",
    "Other": "Other",
    "Misc / Agency": "Other",
}


def get_connection():
    """Get a psycopg2 connection."""
    return psycopg2.connect(DB_URL)


def normalize_industries(conn) -> int:
    """Normalize industry values to canonical names."""
    cur = conn.cursor()
    updated = 0
    for raw, canonical in INDUSTRY_MAP.items():
        if raw == canonical:
            continue
        cur.execute(
            "UPDATE algolia_customers SET industry = %s WHERE industry = %s",
            (canonical, raw),
        )
        if cur.rowcount > 0:
            print(f"  Industry: '{raw}' → '{canonical}' ({cur.rowcount} rows)")
            updated += cur.rowcount
    conn.commit()
    cur.close()
    return updated


def enrich_from_vertical_tabs(conn, wb: openpyxl.Workbook) -> int:
    """Enrich customers with ARR tier and vertical source from vertical tabs."""
    cur = conn.cursor()
    updated = 0

    vertical_tabs = {
        "Grocery": {"industry": "Grocery", "name_col": "account_name", "arr_col": "usd_arr", "website_col": "website"},
        "Fashion": {"industry": "Fashion", "name_col": "Company", "arr_col": "ARR", "website_col": None},
        "Luxury": {"industry": "Luxury", "name_col": "name", "arr_col": "USD_ARR__c", "website_col": "website"},
        "Travel": {"industry": "Travel & Hospitality", "name_col": "Account Name", "arr_col": None, "website_col": "Website"},
        "FinServ": {"industry": "Financial Services", "name_col": "Company", "arr_col": None, "website_col": None},
        "100k+": {"industry": None, "name_col": "Account Name", "arr_col": "ARR", "website_col": "Website"},
    }

    for tab_name, config in vertical_tabs.items():
        ws = wb[tab_name]
        headers = [str(c.value or "").strip() for c in ws[1]]

        # Find column indices
        name_idx = None
        arr_idx = None
        website_idx = None
        for i, h in enumerate(headers):
            if config["name_col"] and h == config["name_col"]:
                name_idx = i
            if config["arr_col"] and h == config["arr_col"]:
                arr_idx = i
            if config["website_col"] and h == config["website_col"]:
                website_idx = i

        if name_idx is None:
            print(f"  WARNING: Could not find name column '{config['name_col']}' in tab '{tab_name}'")
            continue

        tab_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[name_idx] or "").strip()
            if not name:
                continue

            # Determine ARR tier
            arr_tier = None
            if arr_idx is not None and row[arr_idx] is not None:
                try:
                    arr_val = float(str(row[arr_idx]).replace("$", "").replace(",", "").strip())
                    if arr_val >= 100000:
                        arr_tier = "Enterprise 100k+"
                    elif arr_val >= 50000:
                        arr_tier = "Mid-Market 50k+"
                    elif arr_val >= 10000:
                        arr_tier = "Growth 10k+"
                    else:
                        arr_tier = "SMB"
                except (ValueError, TypeError):
                    pass

            # For 100k+ tab, always Enterprise
            if tab_name == "100k+" and arr_tier is None:
                arr_tier = "Enterprise 100k+"

            # Build update
            updates = []
            params = []

            if config["industry"]:
                updates.append("industry = COALESCE(NULLIF(industry, 'Unknown'), %s)")
                params.append(config["industry"])

            if arr_tier:
                updates.append("arr_range = COALESCE(arr_range, %s)")
                params.append(arr_tier)

            updates.append("vertical_source = COALESCE(vertical_source, %s)")
            params.append(tab_name)

            updates.append("updated_at = %s")
            params.append(datetime.now(UTC))

            if not updates:
                continue

            # Match by company name (case-insensitive)
            params.append(name)
            sql = f"UPDATE algolia_customers SET {', '.join(updates)} WHERE LOWER(company_name) = LOWER(%s)"
            cur.execute(sql, params)
            if cur.rowcount > 0:
                tab_count += cur.rowcount

        print(f"  Vertical tab '{tab_name}': enriched {tab_count} customers")
        updated += tab_count

    conn.commit()
    cur.close()
    return updated


def load_adobe_partners(conn, wb: openpyxl.Workbook) -> int:
    """Add Adobe partner cross-reference from Adobe tab."""
    cur = conn.cursor()
    ws = wb["Adobe"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    name_idx = headers.index("Account Name") if "Account Name" in headers else 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_idx] or "").strip()
        if not name:
            continue

        # Add "Adobe" to partner_ecosystem JSONB array if not already present
        cur.execute("""
            UPDATE algolia_customers
            SET partner_ecosystem = CASE
                WHEN partner_ecosystem IS NULL THEN '["Adobe"]'::jsonb
                WHEN NOT partner_ecosystem @> '"Adobe"'::jsonb THEN partner_ecosystem || '"Adobe"'::jsonb
                ELSE partner_ecosystem
            END,
            updated_at = %s
            WHERE LOWER(company_name) = LOWER(%s)
        """, (datetime.now(UTC), name))
        if cur.rowcount > 0:
            updated += cur.rowcount

    conn.commit()
    cur.close()
    print(f"  Adobe partners: tagged {updated} customers")
    return updated


def merge_recommend_quotes(conn, wb: openpyxl.Workbook) -> int:
    """Merge Recommend Quotes tab into algolia_quotes."""
    cur = conn.cursor()
    ws = wb["Recommend Quotes"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    inserted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = dict(zip(headers, row))
        customer = str(vals.get("Customer", "") or "").strip()
        name = str(vals.get("Name", "") or "").strip()
        title = str(vals.get("Title", "") or "").strip()
        quote = str(vals.get("Evidence Type", "") or vals.get("Quote", "") or "").strip()
        industry = str(vals.get("Industry", "") or "").strip()
        country = str(vals.get("Country", "") or "").strip()

        if not customer or not quote:
            continue

        # Check if quote already exists (dedup by customer + quote text start)
        cur.execute(
            "SELECT 1 FROM algolia_quotes WHERE LOWER(customer_name) = LOWER(%s) AND LEFT(quote_text, 50) = LEFT(%s, 50)",
            (customer, quote),
        )
        if cur.fetchone():
            continue

        cur.execute("""
            INSERT INTO algolia_quotes (id, customer_name, person_name, person_title, industry, country, quote_text, evidence_type, source, created_at)
            VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (customer, name, title, industry, country, quote, "Recommend Quote", "CustomerEvidence Excel", datetime.now(UTC)))
        inserted += 1

    conn.commit()
    cur.close()
    print(f"  Recommend Quotes: inserted {inserted} new quotes")
    return inserted


def merge_neural_search_cx(conn, wb: openpyxl.Workbook) -> int:
    """Merge Neural Search CX tab — tag customers as Neural Search users."""
    cur = conn.cursor()
    ws = wb["Neural Search CX"]
    headers = [str(c.value or "").strip() for c in ws[1]]

    account_idx = headers.index("Account") if "Account" in headers else 0
    updated = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[account_idx] or "").strip()
        if not name:
            continue

        # Ensure "Neural Search" is in features_used
        cur.execute("""
            UPDATE algolia_customers
            SET features_used = CASE
                WHEN features_used IS NULL THEN '["Neural Search"]'::jsonb
                WHEN NOT features_used @> '"Neural Search"'::jsonb THEN features_used || '"Neural Search"'::jsonb
                ELSE features_used
            END,
            updated_at = %s
            WHERE LOWER(company_name) = LOWER(%s)
        """, (datetime.now(UTC), name))
        if cur.rowcount > 0:
            updated += cur.rowcount

    conn.commit()
    cur.close()
    print(f"  Neural Search CX: tagged {updated} customers")
    return updated


def print_summary(conn):
    """Print final database state summary."""
    cur = conn.cursor()

    print("\n=== FINAL DATABASE STATE ===")
    tables = [
        ("algolia_customers", None),
        ("algolia_case_studies", None),
        ("algolia_quotes", None),
        ("algolia_proofpoints", None),
        ("algolia_advocates", None),
    ]
    for table, _ in tables:
        cur.execute(f"SELECT count(*) FROM {table}")
        count = cur.fetchone()[0]
        print(f"  {table:30s} {count:>6d} rows")

    # Industry distribution after normalization
    print("\n=== INDUSTRY DISTRIBUTION (top 15) ===")
    cur.execute("SELECT industry, count(*) FROM algolia_customers GROUP BY industry ORDER BY count DESC LIMIT 15")
    for row in cur.fetchall():
        print(f"  {row[0]!s:30s} {row[1]:>5d}")

    # ARR tier distribution
    print("\n=== ARR TIER DISTRIBUTION ===")
    cur.execute("SELECT arr_range, count(*) FROM algolia_customers WHERE arr_range IS NOT NULL GROUP BY arr_range ORDER BY count DESC")
    for row in cur.fetchall():
        print(f"  {row[0]!s:30s} {row[1]:>5d}")

    # Partner ecosystem
    print("\n=== PARTNER TAGS ===")
    cur.execute("SELECT count(*) FROM algolia_customers WHERE partner_ecosystem IS NOT NULL AND partner_ecosystem != '[]'::jsonb")
    print(f"  Customers with partner tags: {cur.fetchone()[0]}")

    # Unknown industries remaining
    cur.execute("SELECT count(*) FROM algolia_customers WHERE industry = 'Unknown' OR industry IS NULL")
    print(f"\n  Remaining 'Unknown' industry: {cur.fetchone()[0]}")

    cur.close()


def main():
    """Run all enrichment steps."""
    print(f"Loading Excel: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        print(f"ERROR: File not found: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    conn = get_connection()

    try:
        print("\n[1/5] Normalizing industries...")
        n1 = normalize_industries(conn)
        print(f"  Total updated: {n1}")

        print("\n[2/5] Enriching from vertical tabs (ARR tier, industry fill)...")
        n2 = enrich_from_vertical_tabs(conn, wb)
        print(f"  Total enriched: {n2}")

        print("\n[3/5] Loading Adobe partner cross-reference...")
        n3 = load_adobe_partners(conn, wb)

        print("\n[4/5] Merging Recommend Quotes...")
        n4 = merge_recommend_quotes(conn, wb)

        print("\n[5/5] Merging Neural Search CX...")
        n5 = merge_neural_search_cx(conn, wb)

        print_summary(conn)

        print(f"\nDone. Total changes: {n1 + n2 + n3 + n4 + n5}")
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    main()
